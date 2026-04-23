"""
Parser for the CCC Estimate Blank XLSM workbook.

HOW TO ACTIVATE PARSING
------------------------
Every field in CELL_MAP currently has None as its cell address — the workbook
was not accessible when this module was written. To activate a field:

  1. Open the workbook in Excel.
  2. Navigate to the "Estimate" sheet (SHEET_NAME below).
  3. Find the cell that holds the input value for that field.
  4. Replace None with the cell address string, e.g. "C8" or "D15".

Fields with None addresses always return None from the parser, which renders
as a yellow "not found in workbook" flag on the import review screen. The user
can correct the value manually before saving.
"""

import logging

import openpyxl

logger = logging.getLogger(__name__)

# Name of the worksheet that contains the estimate inputs.
SHEET_NAME = "Estimate"

# ---------------------------------------------------------------------------
# Cell map: field_name -> cell_address_or_None
# ---------------------------------------------------------------------------
# All addresses are None until confirmed from the actual workbook.
# Replace None with a cell address string (e.g. "C8") to activate parsing.
CELL_MAP: dict[str, str | None] = {
    # ── Project Parameters ─────────────────────────────────────────────────
    "deck_area_sf":       None,   # total deck area, square feet
    "blast_level":        None,   # blast spec string, e.g. "SP-5"
    "abrasive_type":      None,   # abrasive media description
    "abrasive_lb_per_sf": None,   # abrasive consumption, lb/SF

    # ── Materials — Primer ─────────────────────────────────────────────────
    "primer_vol_pct":     None,   # volume solids, %
    "primer_mils":        None,   # dry film thickness, mils
    "primer_gal":         None,   # gallons (may be formula-derived)

    # ── Materials — 2nd Primer ─────────────────────────────────────────────
    "second_primer_vol_pct":  None,
    "second_primer_mils":     None,
    "second_primer_gal":      None,

    # ── Materials — Stripe Prime ───────────────────────────────────────────
    "stripe_prime_vol_pct":   None,
    "stripe_prime_mils":      None,
    "stripe_prime_gal":       None,

    # ── Materials — Stripe Intermediate ───────────────────────────────────
    "stripe_intermediate_vol_pct":  None,
    "stripe_intermediate_mils":     None,
    "stripe_intermediate_gal":      None,

    # ── Materials — Intermediate ───────────────────────────────────────────
    "intermediate_vol_pct":   None,
    "intermediate_mils":      None,
    "intermediate_gal":       None,

    # ── Materials — Finish ─────────────────────────────────────────────────
    "finish_vol_pct":     None,
    "finish_mils":        None,
    "finish_gal":         None,

    # ── Labor — Mobilize ──────────────────────────────────────────────────
    "mobilize_sf_per_hr":          None,
    "mobilize_workers_per_nozzle": None,
    "mobilize_hrs_per_day":        None,
    "mobilize_days":               None,

    # ── Labor — Equip Setup ───────────────────────────────────────────────
    "equip_setup_sf_per_hr":          None,
    "equip_setup_workers_per_nozzle": None,
    "equip_setup_hrs_per_day":        None,
    "equip_setup_days":               None,

    # ── Labor — Scaffold ──────────────────────────────────────────────────
    "scaffold_sf_per_hr":          None,
    "scaffold_workers_per_nozzle": None,
    "scaffold_hrs_per_day":        None,
    "scaffold_days":               None,

    # ── Labor — Containment ───────────────────────────────────────────────
    "containment_sf_per_hr":          None,
    "containment_workers_per_nozzle": None,

    # ── Labor — Masking ───────────────────────────────────────────────────
    "masking_sf_per_hr":          None,
    "masking_workers_per_nozzle": None,

    # ── Labor — Pressure Wash ─────────────────────────────────────────────
    "pressure_wash_sf_per_hr":          None,
    "pressure_wash_workers_per_nozzle": None,

    # ── Labor — Caulking (LF-based in practice; treated as SF until refined)
    "caulking_sf_per_hr":          None,
    "caulking_workers_per_nozzle": None,

    # ── Labor — Blast ─────────────────────────────────────────────────────
    "blast_sf_per_hr":          None,
    "blast_workers_per_nozzle": None,

    # ── Labor — Primer application ────────────────────────────────────────
    "primer_labor_sf_per_hr":          None,
    "primer_labor_workers_per_nozzle": None,

    # ── Labor — 2nd Primer application ───────────────────────────────────
    "second_primer_labor_sf_per_hr":          None,
    "second_primer_labor_workers_per_nozzle": None,

    # ── Labor — Stripe Prime application ─────────────────────────────────
    "stripe_prime_labor_sf_per_hr":          None,
    "stripe_prime_labor_workers_per_nozzle": None,

    # ── Labor — Stripe Intermediate application ───────────────────────────
    "stripe_intermediate_labor_sf_per_hr":          None,
    "stripe_intermediate_labor_workers_per_nozzle": None,

    # ── Labor — Intermediate application ─────────────────────────────────
    "intermediate_labor_sf_per_hr":          None,
    "intermediate_labor_workers_per_nozzle": None,

    # ── Labor — Finish application ────────────────────────────────────────
    "finish_labor_sf_per_hr":          None,
    "finish_labor_workers_per_nozzle": None,

    # ── Labor — Traffic Control ───────────────────────────────────────────
    "traffic_control_sf_per_hr":          None,
    "traffic_control_workers_per_nozzle": None,
    "traffic_control_hrs_per_day":        None,
    "traffic_control_days":               None,

    # ── Labor — Inspection / Touchup ─────────────────────────────────────
    "inspection_touchup_sf_per_hr":          None,
    "inspection_touchup_workers_per_nozzle": None,
    "inspection_touchup_hrs_per_day":        None,
    "inspection_touchup_days":               None,

    # ── Labor — OSHA Training ─────────────────────────────────────────────
    "osha_training_sf_per_hr":          None,
    "osha_training_workers_per_nozzle": None,
    "osha_training_hrs_per_day":        None,
    "osha_training_days":               None,

    # ── Shift Structure ───────────────────────────────────────────────────
    "shift_hours_per_day": None,
    "shift_days_total":    None,
    "crew_size":           None,
    "shifts_per_day":      None,
}

# Fields that hold integer values (crew count, shifts per day).
_INT_FIELDS = {"crew_size", "shifts_per_day"}

# Fields that hold free-text strings rather than numbers.
_STR_FIELDS = {"blast_level", "abrasive_type"}

# Numeric fields that need 3 decimal places.
_THREE_DP_FIELDS = {"abrasive_lb_per_sf"}


def _read_cell(ws, addr: str):
    """Read one cell from the worksheet; return None for empty or error values."""
    try:
        val = ws[addr].value
        if val is None:
            return None
        if isinstance(val, str):
            val = val.strip()
            # openpyxl returns formula-error strings like '#REF!', '#VALUE!'
            if not val or val.startswith("#"):
                return None
        return val
    except Exception as exc:
        logger.warning("Could not read cell %s: %s", addr, exc)
        return None


def _coerce(field: str, raw):
    """Cast a raw cell value to the correct Python type for this field."""
    if raw is None:
        return None

    if field in _STR_FIELDS:
        return str(raw).strip() or None

    if field in _INT_FIELDS:
        try:
            return int(round(float(raw)))
        except (ValueError, TypeError):
            logger.warning("Non-integer value for %s: %r", field, raw)
            return None

    # Default: numeric
    scale = 3 if field in _THREE_DP_FIELDS else 2
    try:
        return round(float(raw), scale)
    except (ValueError, TypeError):
        logger.warning("Non-numeric value for %s: %r", field, raw)
        return None


def parse_estimate_workbook(file_stream) -> dict:
    """
    Parse a CCC Estimate Blank XLSM workbook from a file-like object.

    Returns a dict {field_name: value_or_None} for every field in CELL_MAP.
    Fields whose address is None in CELL_MAP always return None.
    Any cell that cannot be parsed returns None and logs a warning.
    Raises openpyxl exceptions if the file cannot be opened at all.
    """
    result: dict = {field: None for field in CELL_MAP}

    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True, keep_vba=False)

    if SHEET_NAME not in wb.sheetnames:
        logger.warning(
            "Sheet %r not found in workbook. Available sheets: %s",
            SHEET_NAME,
            wb.sheetnames,
        )
        wb.close()
        return result

    ws = wb[SHEET_NAME]

    for field, addr in CELL_MAP.items():
        if addr is None:
            continue
        raw = _read_cell(ws, addr)
        result[field] = _coerce(field, raw)

    logger.info(
        "Parsed %d/%d fields from sheet %r",
        sum(1 for v in result.values() if v is not None),
        len(result),
        SHEET_NAME,
    )

    wb.close()
    return result
